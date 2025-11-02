'''

MIT License

Copyright (c) 2025 Vidyut Prabakaran

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.

'''

## Imports and stuff

import xlsxwriter
import openpyxl
import os
import sys

## Main 1

print ("Welcome to Xcel CLI!")
print(" - When entering filename, you can enter an existing file to edit it, or a new file to create it.")
print(" - Type 'view()' at any time to view the current spreadsheet contents.")
print(" - Type 'exit()' at any time to save and exit the program.")
print(" - Type 'about()' at any time to view program information.\n")

filename = input("File name with .xlsx extension (e.g. Hello.xlsx): ")
wb = xlsxwriter.Workbook(filename)
ws = wb.add_worksheet()
print(f"Spreadsheet '{filename}'! Opening editor...")

## Functions
def about():
    print("\nXcel CLI - A simple and lightweight command line interface for creating, editing, and viewing Excel spreadsheets.")
    print("Created by Vidyut Prabakaran | https://VidyutPrabakaran.github.io ")
    print("Version 1.0\n")

def cmd(cell, value):
    ws.write(cell, value)
    print(f"Written '{value}' to cell '{cell}' successfully!")

def view():
    wb_r = openpyxl.load_workbook(filename)
    ws_r = wb_r.active

    max_col = ws_r.max_column

    for row in ws_r.iter_rows(min_row=1, max_row=ws_r.max_row, max_col=max_col):    # Read all the written rows & columns
        out_row = []
        for cell in row:
            if cell.value:  # If cell is not empty
                out_row.append(str(cell.value))
            else:   # If cell is empty
                out_row.append(" " * 4)   # Add 4 spaces for blank cell
        print(out_row)

## Main 2

def loop():
    i = 0
    while i < 1:
        cell = input("\nEnter cell (e.g. A1): ")
        if cell.lower() == "exit()":
            i += 1
            print(f"\nSaving and exiting '{filename}'...")
            wb.close()
            break

        elif cell.lower() == "view()":
            view()

        elif cell.lower() == "about()":
            about()

        else:
            pass

        value = input("Enter value: ")
        if value.lower() == "exit()":
            i += 1
            print(f"\nSaving and exiting '{filename}'...")
            wb.close()
            break

        elif value.lower() == "view()":
            view()

        elif value.lower() == "about()":
            about()

        else:
            pass

        cmd(cell, value)

loop()